#!/usr/bin/env python3
"""
edgerag.analysis.base

EdgeRAG Phase 1 analysis with:
- config-driven nicknames
- publication figures
- publication tables exported to Excel (.xlsx) and PNG
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import matplotlib
matplotlib.rcParams["pdf.fonttype"] = 42
matplotlib.rcParams["ps.fonttype"] = 42
matplotlib.rcParams["svg.fonttype"] = "path"

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from edgerag.analysis.common import BASE_ANALYSIS_OUTDIR, resolve_config_path, resolve_results_path
from edgerag.core.config import load_config
from edgerag.core.metrics import (
    compute_exact_match,
    compute_f1,
    normalize_text,
    strip_think_prefix_for_scoring,
)

def maybe_rescore_think_row(row: Dict[str, Any]) -> bool:
    """Recompute EM/F1 (and grounded variants) if the stored answer contains a think block."""

    answer = row.get("answer")
    gold_answers = row.get("gold_answers")
    if not isinstance(answer, str) or "</think>" not in answer.lower() or not isinstance(gold_answers, list):
        return False

    scored_answer = strip_think_prefix_for_scoring(answer)
    if scored_answer == answer:
        return False

    em = compute_exact_match(scored_answer, gold_answers)
    f1 = compute_f1(scored_answer, gold_answers)
    row["metric_em"] = em
    row["metric_f1"] = f1

    retrieved_ids = row.get("retrieved_ids") if isinstance(row.get("retrieved_ids"), list) else []
    gold_provenance = row.get("gold_provenance") if isinstance(row.get("gold_provenance"), list) else []
    gold_set = set(gold_provenance)
    retrieved_set = set(retrieved_ids)
    prov_hit = 1.0 if gold_set and gold_set.intersection(retrieved_set) else 0.0
    prov_all = 1.0 if gold_set and gold_set.issubset(retrieved_set) else 0.0
    row["metric_kilt_em_hit"] = float(em) * prov_hit
    row["metric_kilt_f1_hit"] = float(f1) * prov_hit
    row["metric_kilt_em_all"] = float(em) * prov_all
    row["metric_kilt_f1_all"] = float(f1) * prov_all
    return True


# -----------------------------
# Loading
# -----------------------------
def load_results(results_path: Path) -> pd.DataFrame:
    rows = []
    rescored_rows = 0
    with results_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            row = {k: v for k, v in r.items() if k not in ("metrics", "timings")}
            for k, v in r.get("metrics", {}).items():
                row[f"metric_{k}"] = v
            for k, v in r.get("timings", {}).items():
                row[f"time_{k}"] = v
            if maybe_rescore_think_row(row):
                rescored_rows += 1
            rows.append(row)
    if rescored_rows:
        print(f"[load] Re-scored {rescored_rows} rows after stripping leading <think>...</think> content")
    return pd.DataFrame(rows)



def ensure_columns(df: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    df = df.copy()
    for col in columns:
        if col not in df.columns:
            df[col] = np.nan
    return df


# -----------------------------
# Ordering helpers
# -----------------------------
def ordered_present(values: Iterable[str], preferred_order: Iterable[str]) -> List[str]:
    values_set = set(values)
    ordered: List[str] = []
    seen: set[str] = set()

    for v in preferred_order:
        if v in values_set and v not in seen:
            ordered.append(v)
            seen.add(v)

    remainder = sorted(v for v in values_set if v not in seen)
    return ordered + remainder


# -----------------------------
# Model metadata helpers
# -----------------------------
def _coerce_model_list(raw: Any) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    if not isinstance(raw, list):
        return out

    for item in raw:
        if isinstance(item, str):
            out.append({"name": item, "nickname": item})
        elif isinstance(item, dict) and item.get("name"):
            name = str(item["name"])
            nickname = str(item.get("nickname") or item.get("display_name") or item.get("short_name") or name)
            out.append({"name": name, "nickname": nickname})
    return out


def build_model_metadata(cfg: Dict[str, Any], df: pd.DataFrame) -> Dict[str, Any]:
    embed_specs = _coerce_model_list(cfg.get("embed_models", cfg.get("embedders", [])))
    gen_specs = _coerce_model_list(cfg.get("generator_models", cfg.get("generators", [])))

    embed_name_to_label = {spec["name"]: spec["nickname"] for spec in embed_specs}
    gen_name_to_label = {spec["name"]: spec["nickname"] for spec in gen_specs}

    # Common alias for historical HF-prefixed MiniLM rows.
    if "all-MiniLM-L6-v2" in embed_name_to_label:
        embed_name_to_label.setdefault(
            "hf:sentence-transformers/all-MiniLM-L6-v2",
            embed_name_to_label["all-MiniLM-L6-v2"],
        )

    embed_order_raw = [spec["name"] for spec in embed_specs]
    gen_order_raw = [spec["name"] for spec in gen_specs]

    if "embedder" in df.columns:
        seen_embs = [e for e in df["embedder"].dropna().astype(str).unique().tolist() if e != "__none__"]
        for emb in seen_embs:
            embed_name_to_label.setdefault(emb, emb)
        for emb in seen_embs:
            if emb not in embed_order_raw:
                embed_order_raw.append(emb)

    if "generator" in df.columns:
        seen_gens = [g for g in df["generator"].dropna().astype(str).unique().tolist() if g != "__shared_p0__"]
        for gen in seen_gens:
            gen_name_to_label.setdefault(gen, gen)
        for gen in seen_gens:
            if gen not in gen_order_raw:
                gen_order_raw.append(gen)

    embed_order_label = ordered_present(
        [embed_name_to_label.get(name, name) for name in embed_order_raw],
        [embed_name_to_label.get(name, name) for name in embed_order_raw],
    )
    gen_order_label = ordered_present(
        [gen_name_to_label.get(name, name) for name in gen_order_raw],
        [gen_name_to_label.get(name, name) for name in gen_order_raw],
    )

    return {
        "embed_name_to_label": embed_name_to_label,
        "gen_name_to_label": gen_name_to_label,
        "embed_order_raw": embed_order_raw,
        "gen_order_raw": gen_order_raw,
        "embed_order_label": embed_order_label,
        "gen_order_label": gen_order_label,
    }


def label_embedder(embedder: Any, meta: Dict[str, Any]) -> Any:
    if pd.isna(embedder):
        return embedder
    return meta["embed_name_to_label"].get(str(embedder), str(embedder))


def label_generator(generator: Any, meta: Dict[str, Any]) -> Any:
    if pd.isna(generator):
        return generator
    return meta["gen_name_to_label"].get(str(generator), str(generator))


# -----------------------------
# Canonicalization
# -----------------------------
def canonicalize(df: pd.DataFrame, meta: Dict[str, Any]) -> pd.DataFrame:
    df = df.copy()
    df["embedder_label"] = df["embedder"].map(lambda x: label_embedder(x, meta))
    df["generator_label"] = df["generator"].map(lambda x: label_generator(x, meta))
    df["status_norm"] = df["status"].fillna("ok")

    p0 = df[(df["pipeline"] == "P0") & (df["generator"] == "__shared_p0__")].copy()
    p1 = df[(df["pipeline"] == "P1") & (df["embedder"] == "__none__")].copy()

    p23 = df[df["pipeline"].isin(["P2", "P3"])].copy()
    p23["ts"] = pd.to_datetime(p23["timestamp"])
    dedup_key = [
        "pipeline",
        "generator",
        "embedder_label",
        "top_k",
        "context_length",
        "question_id",
    ]
    p23 = p23.sort_values("ts").drop_duplicates(subset=dedup_key, keep="last")

    canon = pd.concat([p0, p1, p23], ignore_index=True)
    canon["status_norm"] = canon["status_norm"].fillna("ok")
    return canon


# -----------------------------
# Summary tables
# -----------------------------
def build_summary_tables(canon: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    p0_summary = (
        canon[canon["pipeline"] == "P0"]
        .groupby(["embedder", "embedder_label", "top_k"], as_index=False)
        .agg(
            n=("question_id", "count"),
            recall_at_k=("metric_recall_at_k", "mean"),
            r_precision=("metric_r_precision", "mean"),
            retrieve_time_s=("time_retrieve_s", "mean"),
        )
        .sort_values(["embedder_label", "top_k"])
    )

    p1_summary = (
        canon[canon["pipeline"] == "P1"]
        .groupby(["generator", "generator_label"], as_index=False)
        .agg(
            n=("question_id", "count"),
            em=("metric_em", "mean"),
            f1=("metric_f1", "mean"),
            total_time_s=("time_total_s", "mean"),
            fail_rate=("status_norm", lambda s: (s == "failed").mean()),
        )
        .sort_values("generator_label")
    )

    p23_cfg = (
        canon[canon["pipeline"].isin(["P2", "P3"])]
        .groupby(["pipeline", "generator", "generator_label", "embedder", "embedder_label", "top_k"], as_index=False)
        .agg(
            n=("question_id", "count"),
            em=("metric_em", "mean"),
            f1=("metric_f1", "mean"),
            recall_at_k=("metric_recall_at_k", "mean"),
            r_precision=("metric_r_precision", "mean"),
            grounded_em_hit=("metric_kilt_em_hit", "mean"),
            grounded_f1_hit=("metric_kilt_f1_hit", "mean"),
            grounded_em_all=("metric_kilt_em_all", "mean"),
            grounded_f1_all=("metric_kilt_f1_all", "mean"),
            total_time_s=("time_total_s", "mean"),
            fail_rate=("status_norm", lambda s: (s == "failed").mean()),
        )
        .sort_values(["pipeline", "generator_label", "f1"], ascending=[True, True, False])
    )

    best_p2 = (
        p23_cfg[p23_cfg["pipeline"] == "P2"]
        .sort_values(["generator", "f1"], ascending=[True, False])
        .groupby("generator", as_index=False)
        .head(1)
        .reset_index(drop=True)
    )

    best_p3 = (
        p23_cfg[p23_cfg["pipeline"] == "P3"]
        .sort_values(["generator", "f1"], ascending=[True, False])
        .groupby("generator", as_index=False)
        .head(1)
        .reset_index(drop=True)
    )

    fail_summary = (
        canon[canon["generator"] != "__shared_p0__"]
        .groupby(["generator", "generator_label", "pipeline"], as_index=False)
        .agg(fail_rate=("status_norm", lambda s: (s == "failed").mean()))
        .sort_values(["generator_label", "pipeline"])
    )

    fail_stage_summary = pd.DataFrame()
    if "error_stage" in canon.columns:
        fail_stage_summary = (
            canon[(canon["status_norm"] == "failed") & (canon["generator"] != "__shared_p0__")]
            .groupby(["pipeline", "error_stage"], as_index=False)
            .agg(n=("question_id", "count"))
            .sort_values(["pipeline", "n"], ascending=[True, False])
        )

    return {
        "p0_summary": p0_summary,
        "p1_summary": p1_summary,
        "p23_by_config": p23_cfg,
        "best_p2": best_p2,
        "best_p3": best_p3,
        "fail_summary": fail_summary,
        "fail_stage_summary": fail_stage_summary,
    }


# -----------------------------
# Publication tables (xlsx + png)
# -----------------------------
def _best_metric_by_generator_topk(canon: pd.DataFrame, pipeline: str, metric_col: str) -> pd.DataFrame:
    sub = (
        canon[canon["pipeline"] == pipeline]
        .groupby(["generator_label", "top_k", "embedder_label"], as_index=False)
        .agg(metric=(metric_col, "mean"))
        .sort_values(["generator_label", "top_k", "metric"], ascending=[True, True, False])
        .drop_duplicates(["generator_label", "top_k"], keep="first")
    )
    return sub


def _build_generator_metric_table(
    canon: pd.DataFrame,
    generator_order: List[str],
    metric_col: str,
    metric_label: str,
    include_p1: bool = True,
) -> pd.DataFrame:
    table = pd.DataFrame({"Generator": generator_order})

    if include_p1:
        p1_metric = (
            canon[canon["pipeline"] == "P1"]
            .groupby("generator_label", as_index=False)
            .agg(value=(metric_col, "mean"))
            .rename(columns={"value": f"P1 {metric_label}"})
        )
        table = table.merge(p1_metric.rename(columns={"generator_label": "Generator"}), on="Generator", how="left")

    for pipeline in ["P2", "P3"]:
        pivot = _best_metric_by_generator_topk(canon, pipeline, metric_col).pivot(
            index="generator_label", columns="top_k", values="metric"
        )
        for k in sorted(k for k in pivot.columns.tolist() if pd.notna(k)):
            table[f"{pipeline} {metric_label} k = {int(k)}"] = table["Generator"].map(pivot[k])

    return table


def build_publication_tables(canon: pd.DataFrame, meta: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    embedder_order = ordered_present(
        canon.loc[canon["pipeline"] == "P0", "embedder_label"].dropna().astype(str).unique().tolist(),
        meta["embed_order_label"],
    )
    generator_order = ordered_present(
        canon.loc[canon["generator"] != "__shared_p0__", "generator_label"].dropna().astype(str).unique().tolist(),
        meta["gen_order_label"],
    )

    retrieval_table = (
        canon[canon["pipeline"] == "P0"]
        .groupby(["embedder_label", "top_k"], as_index=False)
        .agg(
            recall_at_k=("metric_recall_at_k", "mean"),
            r_precision=("metric_r_precision", "mean"),
            retrieve_time_s=("time_retrieve_s", "mean"),
        )
    )
    retrieval_table["Recall@k (%)"] = retrieval_table["recall_at_k"] * 100
    retrieval_table["R-Prec (%)"] = retrieval_table["r_precision"] * 100
    retrieval_table["Mean retrieval time (s)"] = retrieval_table["retrieve_time_s"]
    retrieval_table = retrieval_table[["embedder_label", "top_k", "Recall@k (%)", "R-Prec (%)", "Mean retrieval time (s)"]]
    retrieval_table = retrieval_table.rename(columns={"embedder_label": "Embedder", "top_k": "Top-k"})
    retrieval_table["Embedder"] = pd.Categorical(retrieval_table["Embedder"], categories=embedder_order, ordered=True)
    retrieval_table = retrieval_table.sort_values(["Embedder", "Top-k"]).reset_index(drop=True)
    retrieval_table["Embedder"] = retrieval_table["Embedder"].astype(str)

    generator_f1_table = _build_generator_metric_table(canon, generator_order, "metric_f1", "F1", include_p1=True)
    generator_em_table = _build_generator_metric_table(canon, generator_order, "metric_em", "EM", include_p1=True)
    generator_grounded_f1_hit_table = _build_generator_metric_table(
        canon, generator_order, "metric_kilt_f1_hit", "gF1-hit", include_p1=False
    )
    generator_grounded_em_hit_table = _build_generator_metric_table(
        canon, generator_order, "metric_kilt_em_hit", "gEM-hit", include_p1=False
    )
    generator_grounded_f1_all_table = _build_generator_metric_table(
        canon, generator_order, "metric_kilt_f1_all", "gF1-all", include_p1=False
    )
    generator_grounded_em_all_table = _build_generator_metric_table(
        canon, generator_order, "metric_kilt_em_all", "gEM-all", include_p1=False
    )

    if "error_stage" in canon.columns:
        fail_stage_table = (
            canon[(canon["status_norm"] == "failed") & (canon["generator"] != "__shared_p0__")]
            .groupby(["pipeline", "error_stage"], as_index=False)
            .agg(Count=("question_id", "count"))
            #.sort_values(["pipeline", "n"], ascending=[True, False])
        )
        fail_stage_table["Pipeline"] = pd.Categorical(fail_stage_table["pipeline"], categories=["P3", "P2", "P1"], ordered=True)
        fail_stage_table = fail_stage_table.sort_values(["Pipeline", "Count"], ascending=[True, False])
        fail_stage_table = fail_stage_table.rename(columns={"error_stage": "Error stage"})[["Pipeline", "Error stage", "Count"]].reset_index(drop=True)
        fail_stage_table["Pipeline"] = fail_stage_table["Pipeline"].astype(str)
    else:
        fail_stage_table = pd.DataFrame(columns=["Pipeline", "Error stage", "Count"])

    return {
        "table_retrieval_summary": retrieval_table,
        "table_generator_f1": generator_f1_table,
        "table_generator_em": generator_em_table,
        "table_generator_grounded_f1_hit": generator_grounded_f1_hit_table,
        "table_generator_grounded_em_hit": generator_grounded_em_hit_table,
        "table_generator_grounded_f1_all": generator_grounded_f1_all_table,
        "table_generator_grounded_em_all": generator_grounded_em_all_table,
        "table_failure_stages": fail_stage_table,
    }


# -----------------------------
# Figure generation helpers
# -----------------------------
def save_all_formats(fig: plt.Figure, outdir: Path, stem: str) -> None:
    fig.savefig(outdir / f"{stem}.png", dpi=300, bbox_inches="tight")
    fig.savefig(outdir / f"{stem}.svg", format="svg", bbox_inches="tight")
    fig.savefig(outdir / f"{stem}.pdf", format="pdf", bbox_inches="tight")
    plt.close(fig)


def _heatmap_column_order(canon: pd.DataFrame, pipeline: str, meta: Dict[str, Any]) -> List[str]:
    sub = canon[canon["pipeline"] == pipeline]
    embedders = ordered_present(sub["embedder_label"].dropna().astype(str).unique().tolist(), meta["embed_order_label"])
    top_ks = sorted(int(k) for k in sub["top_k"].dropna().unique().tolist())
    return [f"{emb} k={k}" for emb in embedders for k in top_ks]


# -----------------------------
# Figure generation
# -----------------------------
def figure1_p0(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    p0_agg = (
        canon[canon["pipeline"] == "P0"]
        .groupby(["embedder_label", "top_k"], as_index=False)
        .agg(recall=("metric_recall_at_k", "mean"))
    )
    if p0_agg.empty:
        return

    embedder_order = ordered_present(
        p0_agg["embedder_label"].unique().tolist(),
        meta["embed_order_label"],
    )
    top_ks = sorted(int(k) for k in p0_agg["top_k"].dropna().unique().tolist())
    if not top_ks:
        return

    width = 0.8 / max(len(top_ks), 1)
    fig, ax = plt.subplots(figsize=(max(8, 1.8 * len(embedder_order)), 4.5))
    x = np.arange(len(embedder_order))

    for idx, top_k in enumerate(top_ks):
        vals = []
        for e in embedder_order:
            sub = p0_agg[(p0_agg["embedder_label"] == e) & (p0_agg["top_k"] == top_k)]["recall"]
            vals.append(float(sub.iloc[0]) * 100 if not sub.empty else np.nan)
        offset = (idx - (len(top_ks) - 1) / 2) * width
        bars = ax.bar(x + offset, vals, width, label=f"Recall@{top_k}")
        for b in bars:
            h = b.get_height()
            if np.isfinite(h):
                ax.annotate(
                    f"{h:.1f}",
                    xy=(b.get_x() + b.get_width() / 2, h),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                    fontsize=8,
                )

    ax.set_ylabel("Recall (%)")
    ax.set_title("Shared P0 retrieval performance")
    ax.set_xticks(x)
    ax.set_xticklabels(embedder_order)
    ax.set_ylim(0, 105)
    ax.legend()

    save_all_formats(fig, outdir, "fig1_p0_recall")


def figure2_generator_scaling(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    gen_order_raw = ordered_present(
        canon.loc[canon["generator"] != "__shared_p0__", "generator"].dropna().astype(str).unique().tolist(),
        meta["gen_order_raw"],
    )
    if not gen_order_raw:
        return

    p1_f1 = (
        canon[canon["pipeline"] == "P1"]
        .groupby("generator", as_index=False)["metric_f1"]
        .mean()
        .rename(columns={"metric_f1": "p1_f1"})
    )

    best_p2 = (
        canon[canon["pipeline"] == "P2"]
        .groupby(["generator", "embedder_label", "top_k"], as_index=False)
        .agg(f1=("metric_f1", "mean"))
        .sort_values(["generator", "f1"])
        .groupby("generator", as_index=False)
        .tail(1)
        .rename(columns={"f1": "p2_f1"})
    )

    best_p3 = (
        canon[canon["pipeline"] == "P3"]
        .groupby(["generator", "embedder_label", "top_k"], as_index=False)
        .agg(f1=("metric_f1", "mean"))
        .sort_values(["generator", "f1"])
        .groupby("generator", as_index=False)
        .tail(1)
        .rename(columns={"f1": "p3_f1"})
    )

    scale_df = pd.DataFrame({"generator": gen_order_raw})
    scale_df["generator_label"] = scale_df["generator"].map(lambda x: label_generator(x, meta))
    scale_df = scale_df.merge(p1_f1, on="generator", how="left")
    scale_df = scale_df.merge(best_p2[["generator", "p2_f1"]], on="generator", how="left")
    scale_df = scale_df.merge(best_p3[["generator", "p3_f1"]], on="generator", how="left")
    for col in ["p1_f1", "p2_f1", "p3_f1"]:
        scale_df[col] = scale_df[col] * 100

    fig, ax = plt.subplots(figsize=(max(9, 1.3 * len(gen_order_raw)), 4.8))
    x = np.arange(len(gen_order_raw))
    w = 0.24
    bar_specs = [
        ("p1_f1", "P1", -w),
        ("p2_f1", "Best P2", 0),
        ("p3_f1", "Best P3", w),
    ]
    for col, label, offset in bar_specs:
        bars = ax.bar(x + offset, scale_df[col], w, label=label)
        for b in bars:
            h = b.get_height()
            if np.isfinite(h):
                ax.annotate(
                    f"{h:.1f}",
                    xy=(b.get_x() + b.get_width() / 2, h),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                    fontsize=7,
                )

    ax.set_title("Generator scaling: closed-book vs. best RAG F1")
    ax.set_ylabel("F1 (%)")
    ax.set_xticks(x)
    ax.set_xticklabels(scale_df["generator_label"], rotation=20)
    ax.legend()

    save_all_formats(fig, outdir, "fig2_generator_scaling_f1")


def _figure_heatmap(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any], pipeline: str, stem: str, title: str) -> None:
    gen_order_raw = ordered_present(
        canon.loc[canon["pipeline"] == pipeline, "generator"].dropna().astype(str).unique().tolist(),
        meta["gen_order_raw"],
    )
    gen_order_label = [label_generator(g, meta) for g in gen_order_raw]
    col_order = _heatmap_column_order(canon, pipeline, meta)
    if not gen_order_raw or not col_order:
        return

    heat_df = (
        canon[canon["pipeline"] == pipeline]
        .groupby(["generator", "embedder_label", "top_k"], as_index=False)
        .agg(f1=("metric_f1", "mean"))
    )
    heat_df["col"] = heat_df["embedder_label"] + " k=" + heat_df["top_k"].astype(str)
    heat = heat_df.pivot(index="generator", columns="col", values="f1").reindex(index=gen_order_raw, columns=col_order) * 100

    fig, ax = plt.subplots(figsize=(max(10, 1.1 * len(col_order)), max(4.8, 0.55 * len(gen_order_raw) + 2.5)))
    im = ax.imshow(heat.values, aspect="auto")
    ax.set_title(title)
    ax.set_xticks(np.arange(len(col_order)))
    ax.set_xticklabels(col_order, rotation=25, ha="right")
    ax.set_yticks(np.arange(len(gen_order_raw)))
    ax.set_yticklabels(gen_order_label)
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label("F1 (%)")

    for i in range(heat.shape[0]):
        for j in range(heat.shape[1]):
            val = heat.iloc[i, j]
            if pd.notna(val):
                ax.text(j, i, f"{val:.1f}", ha="center", va="center", fontsize=8)

    save_all_formats(fig, outdir, stem)


def figure3_p2_heatmap(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    _figure_heatmap(canon, outdir, meta, pipeline="P2", stem="fig3_p2_heatmap", title="P2 F1 by generator, embedder, and top-k")


def figure4_p3_heatmap(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    _figure_heatmap(canon, outdir, meta, pipeline="P3", stem="fig4_p3_heatmap", title="P3 F1 by generator, embedder, and top-k")


def figure5_fail_rate(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    gen_order_raw = ordered_present(
        canon.loc[canon["generator"] != "__shared_p0__", "generator"].dropna().astype(str).unique().tolist(),
        meta["gen_order_raw"],
    )
    if not gen_order_raw:
        return
    gen_order_label = [label_generator(g, meta) for g in gen_order_raw]
    pipe_order = ["P1", "P2", "P3"]

    fail_rate = (
        canon[canon["generator"] != "__shared_p0__"]
        .groupby(["generator", "pipeline"], as_index=False)
        .agg(fail_rate=("status_norm", lambda s: (s == "failed").mean()))
    )

    fig, ax = plt.subplots(figsize=(max(9, 1.3 * len(gen_order_raw)), 4.8))
    x = np.arange(len(gen_order_raw))
    w = 0.24

    for idx, pipe in enumerate(pipe_order):
        sub = fail_rate[fail_rate["pipeline"] == pipe].set_index("generator").reindex(gen_order_raw)
        vals = sub["fail_rate"].fillna(0).values * 100
        bars = ax.bar(x + (idx - 1) * w, vals, width=w, label=pipe)
        for b in bars:
            h = b.get_height()
            if h > 0:
                ax.annotate(
                    f"{h:.1f}",
                    xy=(b.get_x() + b.get_width() / 2, h),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                    fontsize=7,
                )

    ax.set_title("Runtime failure rate by generator and pipeline")
    ax.set_ylabel("Failed trials (%)")
    ax.set_xticks(x)
    ax.set_xticklabels(gen_order_label, rotation=20)
    ax.legend()

    save_all_formats(fig, outdir, "fig5_fail_rate")

def figure5b_fail_count(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    gen_order_raw = ordered_present(
        canon.loc[canon["generator"] != "__shared_p0__", "generator"].dropna().astype(str).unique().tolist(),
        meta["gen_order_raw"],
    )
    if not gen_order_raw:
        return
    gen_order_label = [label_generator(g, meta) for g in gen_order_raw]
    pipe_order = ["P1", "P2", "P3"]

    fail_count = (
        canon[canon["generator"] != "__shared_p0__"]
        .groupby(["generator", "pipeline"], as_index=False)
        .agg(fail_count=("status_norm", lambda s: (s == "failed").sum()))
    )

    fig, ax = plt.subplots(figsize=(max(9, 1.3 * len(gen_order_raw)), 4.8))
    x = np.arange(len(gen_order_raw))
    w = 0.24

    for idx, pipe in enumerate(pipe_order):
        sub = fail_count[fail_count["pipeline"] == pipe].set_index("generator").reindex(gen_order_raw)
        vals = sub["fail_count"].fillna(0).values
        bars = ax.bar(x + (idx - 1) * w, vals, width=w, label=pipe)
        for b in bars:
            h = b.get_height()
            if h > 0:
                ax.annotate(
                    f"{int(h)}",
                    xy=(b.get_x() + b.get_width() / 2, h),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                    fontsize=7,
                )

    ax.set_title("Runtime failure count by generator and pipeline")
    ax.set_ylabel("Failed trials (count)")
    ax.set_xticks(x)
    ax.set_xticklabels(gen_order_label, rotation=20)
    ax.legend()

    save_all_formats(fig, outdir, "fig5b_fail_count")


def figure6_pareto(canon: pd.DataFrame, outdir: Path, meta: Dict[str, Any]) -> None:
    p1_best = (
        canon[canon["pipeline"] == "P1"]
        .groupby(["generator", "generator_label"], as_index=False)
        .agg(f1=("metric_f1", "mean"), total_time=("time_total_s", "mean"))
        .sort_values(["generator", "f1"])
        .groupby("generator", as_index=False)
        .tail(1)
    )
    p1_best["pipeline"] = "P1"

    p2_best = (
        canon[canon["pipeline"] == "P2"]
        .groupby(["generator", "generator_label", "embedder_label", "top_k"], as_index=False)
        .agg(f1=("metric_f1", "mean"), total_time=("time_total_s", "mean"))
        .sort_values(["generator", "f1"])
        .groupby("generator", as_index=False)
        .tail(1)
    )
    p2_best["pipeline"] = "P2"

    p3_best = (
        canon[canon["pipeline"] == "P3"]
        .groupby(["generator", "generator_label", "embedder_label", "top_k"], as_index=False)
        .agg(f1=("metric_f1", "mean"), total_time=("time_total_s", "mean"))
        .sort_values(["generator", "f1"])
        .groupby("generator", as_index=False)
        .tail(1)
    )
    p3_best["pipeline"] = "P3"

    pareto = pd.concat([p1_best, p2_best, p3_best], ignore_index=True)
    if pareto.empty:
        return
    pareto["f1"] *= 100

    fig, ax = plt.subplots(figsize=(8.5, 5))
    for pipeline, marker in [("P1", "^"), ("P2", "o"), ("P3", "s")]:
        sub = pareto[pareto["pipeline"] == pipeline]
        ax.scatter(sub["total_time"], sub["f1"], label=pipeline, marker=marker)
        for _, r in sub.iterrows():
            ax.annotate(r["generator_label"], (r["total_time"], r["f1"]), fontsize=8, xytext=(4, 4), textcoords="offset points")

    ax.set_xlabel("Mean total time per question (s)")
    ax.set_ylabel("Best configuration F1 (%)")
    ax.set_title("Quality-latency trade-off")
    ax.legend()

    save_all_formats(fig, outdir, "fig6_pareto")


# -----------------------------
# Excel + PNG table export
# -----------------------------
def _format_publication_table_for_png(name: str, df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if name == "table_retrieval_summary":
        out["Top-k"] = out["Top-k"].map(lambda x: f"{int(x)}")
        out["Recall@k (%)"] = out["Recall@k (%)"].map(lambda x: f"{x:.1f}")
        out["R-Prec (%)"] = out["R-Prec (%)"].map(lambda x: f"{x:.1f}")
        out["Mean retrieval time (s)"] = out["Mean retrieval time (s)"].map(lambda x: f"{x:.3f}")
    elif name in {
        "table_generator_f1",
        "table_generator_em",
        "table_generator_grounded_f1_hit",
        "table_generator_grounded_em_hit",
        "table_generator_grounded_f1_all",
        "table_generator_grounded_em_all",
    }:
        for col in out.columns[1:]:
            out[col] = out[col].map(lambda x: "" if pd.isna(x) else f"{x:.3f}")
    elif name == "table_failure_stages":
        out["Count"] = out["Count"].map(lambda x: f"{int(x)}")
    return out


def export_publication_tables_xlsx(tables: Dict[str, pd.DataFrame], out_path: Path) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    thin_gray = Side(style="thin", color="B7B7B7")
    bottom_dark = Side(style="thin", color="555555")
    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    for idx, (name, df) in enumerate(tables.items()):
        title_map = {
            "table_retrieval_summary": "Retrieval summary",
            "table_generator_f1": "Generator F1",
            "table_generator_em": "Generator EM",
            "table_generator_grounded_f1_hit": "Grounded F1 hit",
            "table_generator_grounded_em_hit": "Grounded EM hit",
            "table_generator_grounded_f1_all": "Grounded F1 all",
            "table_generator_grounded_em_all": "Grounded EM all",
            "table_failure_stages": "Failure stages",
        }
        ws = wb.create_sheet(title=title_map.get(name, name)[:31], index=idx)
        ws.freeze_panes = "A2"

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=bottom_dark, bottom=bottom_dark)

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                is_numeric = isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value)
                cell.alignment = Alignment(horizontal="center" if is_numeric else "left", vertical="center")
                cell.border = Border(bottom=thin_gray)

                if name == "table_retrieval_summary":
                    if df.columns[col_idx - 1] in {"Recall@k (%)", "R-Prec (%)"}:
                        cell.number_format = "0.0"
                    elif df.columns[col_idx - 1] == "Mean retrieval time (s)":
                        cell.number_format = "0.000"
                elif name in {
                    "table_generator_f1",
                    "table_generator_em",
                    "table_generator_grounded_f1_hit",
                    "table_generator_grounded_em_hit",
                    "table_generator_grounded_f1_all",
                    "table_generator_grounded_em_all",
                }:
                    if col_idx > 1:
                        cell.number_format = "0.000"
                elif name == "table_failure_stages" and df.columns[col_idx - 1] == "Count":
                    cell.number_format = "0"

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].tolist()])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 28)

        ws.sheet_view.showGridLines = False

    wb.save(out_path)


def render_publication_table_png(name: str, df: pd.DataFrame, out_path: Path, figw) -> None:
    render_df = _format_publication_table_for_png(name, df)
    nrows, ncols = render_df.shape
    fig_w = max(6.5, figw * ncols)
    fig_h = max(1.2 + 0.42 * (nrows + 1), 2.0)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")

    table = ax.table(
        cellText=render_df.values,
        colLabels=render_df.columns,
        loc="center",
        cellLoc="center",
        colLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1, 1.25)

    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor("white")
        cell.set_linewidth(0.0)
        if row == 0:
            cell.set_text_props(weight="bold")
            cell.visible_edges = "TB"
            cell.set_edgecolor("black")
            cell.set_linewidth(0.8)
        elif row == nrows:
            cell.visible_edges = "B"
            cell.set_edgecolor("black")
            cell.set_linewidth(0.8)
        else:
            cell.visible_edges = ""

        if col == 0 and row > 0:
            cell._loc = "left"

    fig.savefig(out_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def export_publication_tables_png(tables: Dict[str, pd.DataFrame], outdir: Path) -> None:
    for name, df in tables.items():
        if name in {
                    "table_generator_grounded_f1_hit",
                    "table_generator_grounded_em_hit",
                    "table_generator_grounded_f1_all",
                    "table_generator_grounded_em_all",
                    "table_retrieval_summary",
                }:
            render_publication_table_png(name, df, outdir / f"{name}.png",2)
        else:
            render_publication_table_png(name, df, outdir / f"{name}.png",1.5)


# -----------------------------
# Main
# -----------------------------
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run the base EdgeRAG publication analysis on a results.jsonl file"
    )
    parser.add_argument(
        "--results",
        type=Path,
        default=None,
        help="Results JSONL path. Defaults to <results_dir>/results.jsonl from the config.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=resolve_config_path(None),
        help="Canonical config JSON used for model metadata, nicknames, ordering, and default results_dir.",
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        default=BASE_ANALYSIS_OUTDIR,
        help="Directory for canonical exports, CSV tables, figures, and publication tables.",
    )
    return parser



def main(argv: Optional[list[str]] = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)
    results_path = resolve_results_path(args.results, args.config)
    config_path = resolve_config_path(args.config)

    args.outdir.mkdir(parents=True, exist_ok=True)

    if not results_path.exists():
        raise FileNotFoundError(f"Could not find results file: {results_path}")

    df = load_results(results_path)
    df = ensure_columns(df, [
        "error_stage",
        "metric_recall_at_k",
        "metric_r_precision",
        "metric_em",
        "metric_f1",
        "metric_kilt_em_hit",
        "metric_kilt_f1_hit",
        "metric_kilt_em_all",
        "metric_kilt_f1_all",
        "time_retrieve_s",
        "time_total_s",
    ])
    cfg = load_config(config_path)
    meta = build_model_metadata(cfg, df)
    canon = canonicalize(df, meta)

    canon.to_json(args.outdir / "canonical_results.jsonl", orient="records", lines=True, force_ascii=False)

    tables = build_summary_tables(canon)
    for name, table in tables.items():
        table.to_csv(args.outdir / f"{name}.csv", index=False)

    publication_tables = build_publication_tables(canon, meta)
    for name, table in publication_tables.items():
        table.to_csv(args.outdir / f"{name}.csv", index=False)
    export_publication_tables_xlsx(publication_tables, args.outdir / "publication_tables.xlsx")
    export_publication_tables_png(publication_tables, args.outdir)

    figure1_p0(canon, args.outdir, meta)
    figure2_generator_scaling(canon, args.outdir, meta)
    figure3_p2_heatmap(canon, args.outdir, meta)
    figure4_p3_heatmap(canon, args.outdir, meta)
    figure5_fail_rate(canon, args.outdir, meta)
    figure5b_fail_count(canon, args.outdir, meta)
    figure6_pareto(canon, args.outdir, meta)

    summary_path = args.outdir / "summary.txt"
    with summary_path.open("w", encoding="utf-8") as f:
        f.write(f"Raw rows: {len(df)}\n")
        f.write(f"Canonical rows: {len(canon)}\n")
        f.write(f"Config used: {config_path}\n")
        f.write(f"Results used: {results_path}\n")
        f.write("\nGenerated publication tables:\n")
        for name in publication_tables:
            f.write(f"- {name}.csv\n")
            f.write(f"- {name}.png\n")
        f.write("- publication_tables.xlsx\n")
        f.write("\nGenerated figures:\n")
        for stem in [
            "fig1_p0_recall",
            "fig2_generator_scaling_f1",
            "fig3_p2_heatmap",
            "fig4_p3_heatmap",
            "fig5_fail_rate",
            "fig5b_fail_count",
            "fig6_pareto",
        ]:
            f.write(f"- {stem}.png/.svg/.pdf\n")

    print(f"Analysis complete. Outputs written to: {args.outdir}")


if __name__ == "__main__":
    main()
